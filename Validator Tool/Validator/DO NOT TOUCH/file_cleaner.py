# file_cleaner.py
import os
import glob
import pandas as pd
from openpyxl import load_workbook, Workbook

def clean_sheet(df):
    # Drop the first 8 rows and first 3 columns
    df_cleaned = df.iloc[8:, 3:].reset_index(drop=True)
    
    # Check if there are enough rows to set the header
    if df_cleaned.shape[0] <= 0:
        raise IndexError("Not enough rows in the DataFrame after dropping the first 8 rows.")
    
    # Set the correct header row
    df_cleaned.columns = df_cleaned.iloc[0]
    
    # Drop the duplicate header row
    df_cleaned = df_cleaned.iloc[1:]
    
    # Check if there are enough rows after dropping the header row
    if df_cleaned.shape[0] <= 0:
        raise IndexError("Not enough rows in the DataFrame after setting the header row.")
    
    # Remove leading and trailing whitespaces from column names and convert to proper case
    df_cleaned.columns = df_cleaned.columns.str.strip().str.title()
    
    # Check if "Emp No" column exists
    if "Emp No" not in df_cleaned.columns:
        raise KeyError('The column "Emp No" was not found in the sheet.')
    
    # Check for blank columns before "Emp No" and remove them
    emp_no_index = df_cleaned.columns.get_loc("Emp No")
    for col_index in range(emp_no_index - 1, -1, -1):  # Traverse columns from right to left
        if df_cleaned.iloc[:, col_index].notna().any():  # Check if any data is present in the column
            break  # Stop loop if non-empty column is found
        else:
            df_cleaned.drop(df_cleaned.columns[col_index], axis=1, inplace=True)  # Drop the empty column
    
    return df_cleaned

def process_files(directory_path_, output_directory_path_, file_pattern='*.xlsm', sheets_to_process=None):

    if sheets_to_process is None:
        sheets_to_process = ['Additions', 'Deletions', 'Corrections']

    file_paths = glob.glob(os.path.join(directory_path_, file_pattern))

    for file_path in file_paths:
        file_name = os.path.splitext(os.path.basename(file_path))[0]
        output_file_path = os.path.join(output_directory_path_, file_name + '.xlsx')
        
        book = load_workbook(file_path)
        new_book = Workbook()
        new_book.remove(new_book.active)
        
        cleaned_sheets = {}

        for sheet_name in sheets_to_process:
            if sheet_name in book.sheetnames:
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, dtype=str)
                try:
                    df_cleaned = clean_sheet(df)
                    cleaned_sheets[sheet_name] = df_cleaned
                except (KeyError, IndexError) as e:
                    print(f"Skipping sheet '{sheet_name}' in file '{file_path}' due to an error: {e}")

        if cleaned_sheets:
            with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
                for sheet_name, df_cleaned in cleaned_sheets.items():
                    df_cleaned.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
            
            new_book = load_workbook(output_file_path)
            
            if 'DO NOT TOUCH' in book.sheetnames:
                if 'DO NOT TOUCH' in new_book.sheetnames:
                     del new_book['DO NOT TOUCH']
            
            new_book.save(output_file_path)
            new_book.close()
        else:
            print(f"No valid sheets found in file '{file_path}'. Skipping saving.")
        
        book.close()

    print("Processing complete.")
